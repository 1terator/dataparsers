import asyncio

from openpyxl.workbook import Workbook

from publicalt.service import PublicaltParseService


async def to_excel(workbook: Workbook) -> Workbook:
    workbook_sheet = workbook.active
    workbook_sheet.title = "Ticketsnebext Exhibitors"
    parse_service = PublicaltParseService()
    parsed_data = await parse_service.parse()

    global_step = 1
    workbook_sheet.append(parse_service.get_params_names())

    for i in range(len(parsed_data)):
        local_step = 0
        for j in range(len(parsed_data[i])):
            if not isinstance(parsed_data[i][j], list):
                workbook_sheet.cell(row=i + 1 + global_step, column=j + 1).value = str(parsed_data[i][j])
                continue
            for obj in parsed_data[i][j]:
                workbook_sheet.cell(row=i + 1 + global_step, column=j + 1).value = str(obj)
            local_step = max(local_step, len(parsed_data[i][j]) - 1)
        global_step += local_step

    workbook_sheet.column_dimensions["A"].width = 50
    workbook_sheet.column_dimensions["B"].width = 50
    workbook_sheet.column_dimensions["C"].width = 50
    workbook_sheet.column_dimensions["D"].width = 100

    return workbook


if __name__ == "__main__":
    excel = asyncio.run(to_excel(workbook=Workbook()))
