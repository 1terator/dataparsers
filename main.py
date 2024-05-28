import asyncio
import re
import uuid
from typing import Any, Tuple

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from cantonfair.service import CantonfairParseService
from eccmid.service import EccmidParseService
from firabarcelona.service import FiraBarcelonaParseService
from ifema.service import IFemaParseService
from infosecurity.service import InfoSecurityParseService
from mwcbarcelona.service import MVCBarcelonaParseService
from publicalt.service import PublicaltParseService
from services import WhatsappService
from simaexpo.service import SimaExpoParseService
from ticketsnebext.service import TicketsNebextParseService
from tools import validate_phone_number, recreate_phone_number


def generate_excel_filename(filename: str) -> str:
    return filename + f"_{str(uuid.uuid4())[:6]}.xlsx"


def is_href(text: str) -> bool:
    return True if re.fullmatch(r"(?P<url>https?://[^\s]+)", text) else False


def to_excel(parsed_data_names: Tuple, parsed_data: Any, sheet_name: str) -> Workbook:
    print(parsed_data)
    workbook = Workbook()
    workbook_sheet = workbook.active
    workbook_sheet.title = sheet_name

    for i in range(len(parsed_data_names)):
        workbook_sheet.cell(row=1, column=i + 1).value = str(parsed_data_names[i])
        workbook_sheet.cell(row=1, column=i + 1).font = Font(bold=True)
        workbook_sheet.column_dimensions[get_column_letter(i+1)].width = 50

    global_step = 1
    for i in range(len(parsed_data)):
        local_step = 0
        for j in range(len(parsed_data[i])):
            if not isinstance(parsed_data[i][j], list):
                if not parsed_data[i][j]:
                    continue
                data = ILLEGAL_CHARACTERS_RE.sub(r'', str(parsed_data[i][j]))
                workbook_sheet.cell(row=i + 1 + global_step, column=j + 1).value = data
                if is_href(text=str(parsed_data[i][j])):
                    workbook_sheet.cell(row=i + 1 + global_step, column=j + 1).style = "Hyperlink"
                    workbook_sheet.cell(row=i + 1 + global_step, column=j + 1).hyperlink = data
                continue
            for k in range(len(parsed_data[i][j])):
                if not parsed_data[i][j][k]:
                    continue
                data = ILLEGAL_CHARACTERS_RE.sub(r'', str(parsed_data[i][j][k]))
                workbook_sheet.cell(row=i + k + 1 + global_step, column=j + 1).value = data
                if is_href(text=str(parsed_data[i][j][k])):
                    workbook_sheet.cell(row=i + k + 1 + global_step, column=j + 1).style = "Hyperlink"
                    workbook_sheet.cell(row=i + k + 1 + global_step, column=j + 1).hyperlink = data
            local_step = max(local_step, len(parsed_data[i][j]) - 1)
        global_step += local_step

    return workbook


async def main():
    filename: str = "InfoService2024"

    whatsapp = WhatsappService(7103909222, "0b7c68fbd0284e098b454ef95d925bf43c48b75d0cc14415a7")
    # publicalt = PublicaltParseService("ExpoBeautyBarcelona2024", whatsapp)
    # ticketsnebext = TicketsNebextParseService("advanced_factories_2024", whatsapp)
    # mwcbarcelona = MVCBarcelonaParseService("00422c3d9f3484bccfae011262fcf49a", "8VVB6VR33K", whatsapp)
    # ifema = IFemaParseService("3a88c5e5-a6e1-4898-b72b-103e4eed1731", "1a015dd8-4c05-4192-2715-08db8781d84f", whatsapp)
    # hispack = FiraBarcelonaParseService("J011024", 136, "hispack2024", whatsapp)
    # bridal = FiraBarcelonaParseService("J113024", 137, whatsapp)
    # iotswc24 = FiraBarcelonaParseService("J025024", 138, "construmat2024", whatsapp)
    # eccmid = EccmidParseService(whatsapp)
    # d5cd7d4ec26134ff4a34d736a7f9ad47
    infoservice = InfoSecurityParseService("d5cd7d4ec26134ff4a34d736a7f9ad47", "XD0U5M6Y4R", whatsapp)
    excel = to_excel(
        parsed_data_names=infoservice.get_parsed_data_names(),
        parsed_data=await infoservice.parse(),
        sheet_name=filename
    )
    await asyncio.to_thread(excel.save, filename=generate_excel_filename(filename))


def format_phone_number(phone_number: str) -> str:
    if not phone_number:
        return phone_number
    phone_number: str = str(int("".join(ch for ch in phone_number if ch.isdigit())))
    return "34" + phone_number if not 11 <= len(phone_number) else phone_number


async def phones(filename: str):
    whatsapp = WhatsappService(7103909222, "0b7c68fbd0284e098b454ef95d925bf43c48b75d0cc14415a7")
    phones_list = []

    with open(filename) as my_file:
        for line in my_file:
            whatsapp_link: str = await whatsapp.format_to_whatsapp_link(
                format_phone_number(recreate_phone_number(line))
            )
            phones_list.append(whatsapp_link if whatsapp_link else "")

    with open("results.txt", "w") as my_file:
        my_file.writelines([phone + "\n" for phone in phones_list])


async def test(filename: str):
    whatsapp = WhatsappService(7103909222, "0b7c68fbd0284e098b454ef95d925bf43c48b75d0cc14415a7")
    canton = SimaExpoParseService(whatsapp)
    excel = to_excel(
        parsed_data_names=canton.get_parsed_data_names(),
        parsed_data=await canton.parse(),
        sheet_name=filename
    )
    await asyncio.to_thread(excel.save, filename=generate_excel_filename(filename))


# asyncio.run(main())
# asyncio.run(phones("phones.txt"))
asyncio.run(main())
